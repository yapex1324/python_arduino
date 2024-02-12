import openpyxl
import numpy as np
from scipy.fft import fft, ifft
import matplotlib.pyplot as plt

# Nama file Excel
file_path = 'data.xlsx'

# Membaca data dari file Excel
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

# Mengambil data dari setiap kolom
data_columns = [list(col) for col in sheet.iter_cols(values_only=True)]

# Menerapkan transformasi Fourier dan filtering pada setiap kolom
filtered_data_columns = []

for i, col_data in enumerate(data_columns, start=1):
    # Mengaplikasikan transformasi Fourier
    fft_result_col = fft(col_data)

    # Menghapus komponen frekuensi tinggi dengan zeroing pada frekuensi tertentu
    cutoff_frequency = 0.1  # Frekuensi cutoff yang diinginkan
    fft_result_filtered_col = fft_result_col.copy()
    num_elements = len(fft_result_col)
    cutoff_index = int(cutoff_frequency * num_elements)
    fft_result_filtered_col[cutoff_index:num_elements-cutoff_index] = 0

    # Mengaplikasikan inverse transform untuk mendapatkan data hasil filtering
    filtered_data_col = ifft(fft_result_filtered_col).real

    # Menambahkan data hasil filtering ke dalam list
    filtered_data_columns.append(filtered_data_col)

    # Menyimpan data hasil filtering ke dalam file Excel
    sheet.cell(row=1, column=len(data_columns) + i, value=f'Filtered Col {i}')
    for j, value in enumerate(filtered_data_col, start=2):
        sheet.cell(row=j, column=len(data_columns) + i, value=value)

# Menyimpan hasil perubahan ke dalam file Excel baru
output_file_path = 'filtered_data.xlsx'
wb.save(output_file_path)

# Menampilkan plot data asli dan hasil filtering
plt.figure(figsize=(12, 10))

for i, (col_data, filtered_data_col) in enumerate(zip(data_columns, filtered_data_columns), start=1):
    # Menampilkan plot data kolom
    plt.subplot(len(data_columns), 2, 2*i-1)
    plt.plot(col_data)
    plt.title(f'Data Asli - Kolom {i}')

    # Menampilkan plot data hasil filtering
    plt.subplot(len(data_columns), 2, 2*i)
    plt.plot(filtered_data_col)
    plt.title(f'Data Setelah Filtering - Kolom {i}')

plt.tight_layout()
plt.show()

# Memberi informasi bahwa data telah disimpan ke dalam file Excel baru
print(f"Data hasil filtering disimpan ke dalam file Excel: {output_file_path}")
